# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
# openpyxl.load_workbook(filename) returns a Workbook object.
# get_sheet_names() and get_sheet_by_name() help get Worksheet objects.
# The square brackets in sheet[‘A1'] get Cell objects.
# Cell objects have a "value" member variable with the content of that cell.
# The cell() method also returns a Cell object from a sheet.

# Import relevant modules
import openpyxl, os

# Update Working Directory to Local Lesson Folder where the example.xlsx file resides
os.chdir('C:\\Development\\Python-Projects\\Automate-The-Boring-Stuff-With-Python\\Section_14_Excel_Word_PDF_Documents\\Lesson_42_ReadingExcelSpreadsheets')

# Verify the Current Working Directory
filePath = os.getcwd()

# DEBUG Print the Current Working Directories File Path
print(filePath)

# Create a Workbook Object by passing the File Name IF in the Current Working Directory
exampleWorkbook = openpyxl.load_workbook('example.xlsx')

# Create a Sheet Object from the workbook (New Syntax)
sheet1 = exampleWorkbook['Sheet1']

# Create a List Variable holding the Workbook Sheet Names (New Syntax)
exampleWorkbookSheetNames = exampleWorkbook.sheetnames

# DEBUG Print the exampleWorkbook Sheet Names
print(exampleWorkbookSheetNames)

# DEBUG Print the exampleWorkbook's Sheet1 selection value for the A1 Cell
# print(sheet1['A1'])

# Create Variables for different Sheet1 Cell values
cellSheet1A1 = sheet1['A1']
cellSheet1B1 = sheet1['B1']
cellSheet1C1 = sheet1['C1']

# DEBUG Print the exampleWorkbook's Sheet1 A1 Cell value in various ways
print(cellSheet1A1.value)
print(str(sheet1['A1'].value))
print(sheet1.cell(row=1, column=1))

# DEBUG Print the exampleWorkbook's Sheet1 B1 Cell value in various ways
print(cellSheet1B1.value)
print(str(sheet1['B1'].value))
print(sheet1.cell(row=1, column=2))

# DEBUG Print the exampleWorkbook's Sheet1 C1 Cell value in various ways
print(cellSheet1C1.value)
print(str(sheet1['C1'].value))
print(sheet1.cell(row=1, column=3))

# DEBUG Print the Cell values for the B Column
for i in range(1, 8):
    print(i, sheet1.cell(row=i, column=2).value)
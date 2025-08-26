# You can view and modify a sheet's name with its "title" member variable.
# Changing a cell's value is done using the square brackets, just like changing a value in a list or dictionary.
# Changes you make to the workbook object can be saved with the save() method.

# Import relevant modules
import openpyxl, os

# Update Working Directory to Local Lesson Folder where the example.xlsx file resides
os.chdir('C:\\Development\\Python-Projects\\Automate-The-Boring-Stuff-With-Python\\Section_14_Excel_Word_PDF_Documents\\Lesson_43_EditingExcelSpreadsheets')

# Verify the Current Working Directory
filePath = os.getcwd()

# DEBUG Print the Current Working Directories File Path
print(filePath)

# Create a new Workbook Object to interact with
exampleWorkbook = openpyxl.Workbook()

# Create a Sheet1 Object from the workbook (New Syntax)
sheet1 = exampleWorkbook['Sheet']

# Create a List Variable holding the Workbook Sheet Names (New Syntax)
exampleWorkbookSheetNames = exampleWorkbook.sheetnames

# DEBUG Print the exampleWorkbook Sheet Names
print(exampleWorkbookSheetNames)

# DEBUG Print the exampleWorkbook's Sheet1 selection value for the A1 Cell
print(sheet1['A1'])

# Create Variables for different Sheet1 Cell values
cellSheet1A1 = sheet1['A1']
cellSheet1B1 = sheet1['B1']
cellSheet1C1 = sheet1['C1']

# DEBUG Print the exampleWorkbook's SheSheet1et A1 Cell value in various ways
print(cellSheet1A1.value)
print(str(sheet1['A1'].value))
print(sheet1.cell(row=1, column=1))

# DEBUG Print the exampleWorkbook's Sheet1 B1 Cell value in various ways
print(cellSheet1B1.value)
print(str(sheet1['B1'].value))
print(sheet1.cell(row=1, column=2))

# DEBUG Print the exampleWorkbook's Sheet1 C1 Cell value in various ways
print(cellSheet1C1.value)
print(str(sheet1['C1'].value))
print(sheet1.cell(row=1, column=3))

# Edit the existing exampleWorkbook's Sheet1 Cell values in various ways
sheet1['A1'] = 42                                    # Update to an Integer value
sheet1['B1'] = 'Hello, this is the B1 Cell'          # Update to an Sting value
sheet1['C1'] = True                                  # Update to an Boolean value

# DEBUG Print the exampleWorkbook's Sheet1 A1 Cell value in various ways
print(cellSheet1A1.value)
print(str(sheet1['A1'].value))
print(sheet1.cell(row=1, column=1))

# DEBUG Print the exampleWorkbook's Sheet1 B1 Cell value in various ways
print(cellSheet1B1.value)
print(str(sheet1['B1'].value))
print(sheet1.cell(row=1, column=2))

# DEBUG Print the exampleWorkbook's Sheet1 C1 Cell value in various ways
print(cellSheet1C1.value)
print(str(sheet1['C1'].value))
print(sheet1.cell(row=1, column=3))

# Save the Workbook Object to a Local file in the Current Working Directory
exampleWorkbook.save('example2.xlsx')

# Create a Sheet2 Object for the workbook (New Syntax)
sheet2 = exampleWorkbook.create_sheet()

# Create a List Variable holding the Workbook Sheet Names (New Syntax)
exampleWorkbookSheetNames = exampleWorkbook.sheetnames

# DEBUG Print the exampleWorkbook Sheet Names
print(exampleWorkbookSheetNames)

# Update the Sheetnames
sheet2.title = 'Sheet2'             # Update Sheet2 1st, to avoid duplicate naming
sheet1.title = 'Sheet1'

# Create a List Variable holding the Workbook Sheet Names (New Syntax)
exampleWorkbookSheetNames = exampleWorkbook.sheetnames

# DEBUG Print the exampleWorkbook Sheet Names
print(exampleWorkbookSheetNames)

# Save the Workbook Object to a Local file in the Current Working Directory
exampleWorkbook.save('example2.xlsx')

# Create a Sheet3 Object for the workbook at the beginning of the Workbook
sheet3 = exampleWorkbook.create_sheet(index=0, title='My 3rd Sheet')

# Create a List Variable holding the Workbook Sheet Names (New Syntax)
exampleWorkbookSheetNames = exampleWorkbook.sheetnames

# DEBUG Print the exampleWorkbook Sheet Names
print(exampleWorkbookSheetNames)

# Save the Workbook Object to a Local file in the Current Working Directory
exampleWorkbook.save('example3.xlsx')
